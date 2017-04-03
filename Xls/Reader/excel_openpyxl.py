from openpyxl import load_workbook
import json
import sys
import os
import argparse
import datetime


def run(argv):
  parser = argparse.ArgumentParser()
  parser.add_argument('--size')
  parser.add_argument('--start')
  parser.add_argument('--action')
  parser.add_argument('--file')
  parser.add_argument('--max-empty-rows', dest="max_empty_rows", default=0)
  parser.add_argument('--process-empty-rows', dest="process_empty_rows", default=False)
  args = parser.parse_args()

  if False == os.path.isfile(args.file):
    print("File does not exist")
    sys.exit(1)

  workbook = load_workbook(args.file, read_only=True)
  sheet = workbook.active
  if sheet.max_row == 65536:
     sheet.max_row = None

  if args.action == "count":
    max_count_empty_rows = int(args.max_empty_rows)
    process_empty_rows = bool(args.process_empty_rows)
    rows_count = 0
    empty_rows_count = 0
    for row in sheet.iter_rows(row_offset=int(1) - 1):
      current_row_read = []
      for cell in row:
        value = cell.value
        if value is None:
          continue
        else:
          current_row_read.append(True)
          if process_empty_rows:
            rows_count += empty_rows_count
          empty_rows_count = 0
          break
      if not current_row_read:
        empty_rows_count += 1
      else:
        rows_count += 1
      if max_count_empty_rows < empty_rows_count:
        break
    print(rows_count)
  elif args.action == "read":
    max_count_empty_rows = int(args.max_empty_rows)
    process_empty_rows = bool(args.process_empty_rows)
    empty_rows_count = 0
    rows = []
    for row in sheet.iter_rows(row_offset=int(args.start) - 1):
        current_row_read = []
        for cell in row:
            value = cell.value
            if isinstance(value, (int, long, float)):
                current_row_read.append(str(value))
            elif value is None:
                current_row_read.append("")
            elif isinstance(value, (datetime.datetime)):
                current_row_read.append(value.strftime("%m/%d/%Y"))
            elif isinstance(value, (str, unicode)):
                current_row_read.append(value.encode('utf-8'))
            else:
                raise TypeError(type(value))
        if any(current_row_read) or process_empty_rows:
            rows.append(current_row_read)
        else:
            empty_rows_count += 1
        if len(rows) >= int(args.size) or max_count_empty_rows < empty_rows_count:
          break
    print(json.dumps(rows))
  else:
    print("Unknown command")
    sys.exit(1)

if __name__ == "__main__":
   run(sys.argv[1:])

